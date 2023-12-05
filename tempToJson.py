import temp  
import pandas as pd
import openpyxl as xl


def getPhaseTenure(PhaseName):
    df = pd.read_excel("template.xlsx" , sheet_name= "PhasesDetails")
    for index, row in df.iterrows():
        if row["Phases"] == PhaseName:
            return int(row['Tenure'])
"""
wb = openpyxl.Workbook()
data = temp.merged_dict
sheet = wb.create_sheet(title= "Sheet")
"""


phaseList = []
groupList = {}
phaseTenure = {}
products = []

for phase , pVals in temp.merged_dict.items():
    phaseList.append(phase)
    groupList[phase] = []
    newP = phase.split("_")[0]
    if newP in temp.getPhases():
        phaseTenure[phase] = getPhaseTenure(newP)
    for groups , gVals in pVals.items():
        groupList[phase].append(groups)
        for items, iVals in gVals.items():
            products.append(items)

products = list(set(products))



tblHeaders = ["Phase Name", "Duration" , "Group Name", "Group Qty"]
tblHeaders += products

workbook = xl.Workbook()
worksheet = workbook.active

for col_num, header in enumerate(tblHeaders, 1):
    worksheet.cell(row=1, column=col_num, value=header)

for row_num, row_data in enumerate(phaseList, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        worksheet.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook
workbook.save('crm_template.xlsx')