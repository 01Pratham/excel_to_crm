import tempTable as 
import pandas as pd
import openpyxl as xl




workbook = xl.Workbook()
worksheet = workbook.active

for col_num, header in enumerate(tblHeaders, 1):
    worksheet.cell(row=1, column=col_num, value=header)

for row_num, row_data in enumerate(phaseList, 2):
    for col_num, cell_value in enumerate(row_data, 1):
        worksheet.cell(row=row_num, column=col_num, value=cell_value)

# Save the workbook
workbook.save('../ExcelTemplates/crm_template.xlsx')